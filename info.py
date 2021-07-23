from openpyxl import load_workbook


class ExcelOperate:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.active_sheet = None

    def read_all_rows(self):
        """
        按行读取整个表格
        :return:
        """
        sheet = self.wb.worksheets[0]
        r_num = sheet.max_row
        c_num = sheet.max_row
        if sheet is not None:
            return r_num, c_num, list(sheet.values)

    def write_list_to_column(self, column, element_list):
        """
        把特定list写入表中某列
        :param column:
        :return:
        """
        sheet = self.wb.worksheets[0]
        for i in range(1, len(element_list) + 2):
            if i == 1:
                sheet.cell(1, column, "Description")
            else:
                sheet.cell(i, column, element_list[i - 2])
        self.wb.save(filename=self.file_path)


zero = "0x0000000000000000000000000000000000000000"
founder = ""
f_num = ""
token = ""
token_name = "AIIS"

if __name__ == "__main__":
    eo = ExcelOperate("test.xlsx")
    row_num, column_num, sh = eo.read_all_rows()
    list_info = []
    for mark in range(1, row_num):
        from_user = sh[mark][4]
        to_user = sh[mark][5]
        quantity = sh[mark][6]
        method = sh[mark][7]

        if from_user == zero:  # 创建
            founder = to_user
            f_num = founder[-3:]
            list_info.append("创始人" + f_num + "创建合约，初始化" + str(quantity) + "数量的" + token_name)
        elif method == "Add Liquidity ETH":  # 增加流动性
            token = to_user
            list_info.append("创始人" + f_num + "在uniswap中增加流动性")
        elif method == "Remove Liquidity ETH":  # 减少流动性
            list_info.append("创始人" + f_num + "在uniswap中减少流动性，取回ETH和" + token_name)
        elif method == "Transfer":
            list_info.append(token_name + "内部转账")
        elif from_user == token and to_user != founder and (
                method == "Swap Exact ETH For Tokens" or method == "Swap ETH For Exact Tokens"):
            list_info.append("受害人交易，ETH交换" + token_name)
        elif from_user == token and to_user == founder and (
                method == "Swap Exact ETH For Tokens" or method == "Swap ETH For Exact Tokens"):
            list_info.append(f_num + "测试交易，ETH交换" + token_name)
        elif from_user == founder and to_user == token and (
                method == "Swap Exact Tokens For ETH" or method == "Swap Tokens For Exact ETH"):
            list_info.append(f_num + "交易，" + token_name + "交换ETH")
        else:
            list_info.append(method)
    eo.write_list_to_column(9, list_info)
    print("done!")
